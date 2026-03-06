#!/usr/bin/env python3
"""Generate review Excel from crawl results in the DB.

Queries hospitals and doctors tables for a given set of place_ids
and creates a multi-sheet Excel workbook for human review.

Usage:
    python3 generate_review.py --place-ids 20951918,1721660349 --output reviews/batch_001.xlsx
    python3 generate_review.py --place-ids 20951918 --db hospitals.db --output review.xlsx
"""

import argparse
import json
import sqlite3
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from storage_manager import DB_DEFAULT


FILL_SUCCESS = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_PARTIAL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_FAILED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def get_db(db_path: str) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def generate_review_excel(db_path: str, place_ids: list, output_path: str) -> str:
    """Generate review Excel for given place_ids. Returns output path."""
    conn = get_db(db_path)
    placeholders = ",".join("?" for _ in place_ids)

    # Query hospitals
    hospitals = conn.execute(
        f"SELECT place_id, csv_no, name, url, final_url, status, "
        f"cms_platform, doctor_page_exists, crawled_at "
        f"FROM hospitals WHERE place_id IN ({placeholders})",
        place_ids,
    ).fetchall()

    # Query doctors with hospital name
    doctors = conn.execute(
        f"SELECT d.place_id, h.name AS hospital_name, d.name AS doctor_name, "
        f"d.role, d.profile_raw_json, d.source_url, d.screenshot_path, "
        f"d.extraction_source, d.ocr_source, d.photo_url "
        f"FROM doctors d JOIN hospitals h ON d.place_id = h.place_id "
        f"WHERE d.place_id IN ({placeholders}) "
        f"ORDER BY d.place_id, d.id",
        place_ids,
    ).fetchall()

    # Doctor count per hospital
    doc_counts = {}
    for d in doctors:
        doc_counts[d["place_id"]] = doc_counts.get(d["place_id"], 0) + 1

    conn.close()

    # Create workbook
    wb = openpyxl.Workbook()

    # --- Sheet 1: Batch Overview ---
    ws1 = wb.active
    ws1.title = "Batch Overview"
    overview_headers = [
        "place_id", "hospital_pk", "name", "website", "final_url",
        "status", "doctor_count", "doctor_page", "cms", "crawled_at",
    ]
    _write_header(ws1, overview_headers)

    for i, h in enumerate(hospitals, start=2):
        row = [
            h["place_id"], h["csv_no"], h["name"], h["url"], h["final_url"],
            h["status"], doc_counts.get(h["place_id"], 0),
            "Yes" if h["doctor_page_exists"] else "No",
            h["cms_platform"] or "", h["crawled_at"] or "",
        ]
        for col, val in enumerate(row, start=1):
            cell = ws1.cell(row=i, column=col, value=val)
        # Status color
        status_cell = ws1.cell(row=i, column=6)
        if h["status"] == "success":
            status_cell.fill = FILL_SUCCESS
        elif h["status"] == "partial":
            status_cell.fill = FILL_PARTIAL
        elif h["status"] in ("failed", "encoding_error"):
            status_cell.fill = FILL_FAILED

    _auto_width(ws1)

    # --- Sheet 2: Doctor Details ---
    ws2 = wb.create_sheet("Doctor Details")
    doc_headers = [
        "place_id", "hospital_name", "doctor_name", "role", "credentials",
        "source_url", "screenshot_path", "extraction", "ocr", "photo_url",
    ]
    _write_header(ws2, doc_headers)

    for i, d in enumerate(doctors, start=2):
        credentials = _format_credentials(d["profile_raw_json"])
        row = [
            d["place_id"], d["hospital_name"], d["doctor_name"], d["role"],
            credentials, d["source_url"] or "", d["screenshot_path"] or "",
            d["extraction_source"] or "", "Yes" if d["ocr_source"] else "No",
            d["photo_url"] or "",
        ]
        for col, val in enumerate(row, start=1):
            cell = ws2.cell(row=i, column=col, value=val)
            if col == 5:  # credentials - wrap text
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    _auto_width(ws2)

    # --- Sheet 3: Not Found (hospitals with 0 doctors) ---
    no_doc_hospitals = [h for h in hospitals if doc_counts.get(h["place_id"], 0) == 0]
    if no_doc_hospitals:
        ws3 = wb.create_sheet("No Doctors Found")
        nd_headers = ["place_id", "hospital_pk", "name", "website", "status", "cms"]
        _write_header(ws3, nd_headers)
        for i, h in enumerate(no_doc_hospitals, start=2):
            row = [h["place_id"], h["csv_no"], h["name"], h["url"], h["status"], h["cms_platform"] or ""]
            for col, val in enumerate(row, start=1):
                ws3.cell(row=i, column=col, value=val)
        _auto_width(ws3)

    # Save
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()

    return output_path


def _write_header(ws, headers: list):
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"


def _auto_width(ws, max_width: int = 50):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val = str(cell.value) if cell.value else ""
            # Use first line only for width calc
            first_line = val.split("\n")[0] if "\n" in val else val
            max_len = max(max_len, len(first_line))
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def _format_credentials(raw_json: str) -> str:
    """Format profile_raw_json array into readable multi-line string."""
    if not raw_json:
        return ""
    try:
        items = json.loads(raw_json)
        if isinstance(items, list):
            return "\n".join(str(item) for item in items)
        return str(items)
    except (json.JSONDecodeError, TypeError):
        return str(raw_json)


def main():
    parser = argparse.ArgumentParser(description="Generate review Excel from crawl DB")
    parser.add_argument("--db", default=DB_DEFAULT, help=f"SQLite DB path (default: {DB_DEFAULT})")
    parser.add_argument("--place-ids", required=True, help="Comma-separated place IDs")
    parser.add_argument("--output", required=True, help="Output Excel path")
    args = parser.parse_args()

    place_ids = [pid.strip() for pid in args.place_ids.split(",")]
    output = generate_review_excel(args.db, place_ids, args.output)
    print(f"Review Excel generated: {output} ({len(place_ids)} hospitals)")


if __name__ == "__main__":
    main()

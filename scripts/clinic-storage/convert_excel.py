#!/usr/bin/env python3
"""Convert hospitals_for_crawl.xlsx to crawler-compatible CSV.

Reads the specified sheet from the Excel file and maps columns:
  id → csv_no, name → naver_name, naver_place_id → place_id, website → homepage_url

Usage:
    python3 convert_excel.py --xlsx data/hospitals_for_crawl.xlsx --sheet "Real Website"
    python3 convert_excel.py --xlsx data/hospitals_for_crawl.xlsx --sheet "Social Channel" --output social.csv
"""

import argparse
import csv
import sys
from pathlib import Path

import openpyxl


def convert(xlsx_path: str, sheet_name: str, output_path: str) -> int:
    """Convert Excel sheet to crawler CSV. Returns row count."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    col_map = {str(h).strip(): i for i, h in enumerate(header)}

    required = {"id", "name", "naver_place_id", "website"}
    missing = required - set(col_map.keys())
    if missing:
        print(f"Error: Missing columns: {missing}", file=sys.stderr)
        sys.exit(1)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    count = 0
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=["csv_no", "place_id", "naver_name", "homepage_url"])
        writer.writeheader()
        for row in rows[1:]:
            place_id = row[col_map["naver_place_id"]]
            if place_id is None:
                continue
            if isinstance(place_id, float):
                place_id = str(int(place_id))
            else:
                place_id = str(place_id).strip()

            website = row[col_map["website"]]
            if not website or not str(website).strip():
                continue

            writer.writerow({
                "csv_no": row[col_map["id"]],
                "place_id": place_id,
                "naver_name": row[col_map["name"]],
                "homepage_url": str(website).strip(),
            })
            count += 1

    wb.close()
    return count


def main():
    parser = argparse.ArgumentParser(description="Convert Excel to crawler CSV")
    parser.add_argument("--xlsx", required=True, help="Path to hospitals_for_crawl.xlsx")
    parser.add_argument("--sheet", default="Real Website", help="Sheet name (default: Real Website)")
    parser.add_argument("--output", default=None, help="Output CSV path (default: auto)")
    args = parser.parse_args()

    if not args.output:
        project_root = Path(__file__).resolve().parent.parent.parent
        args.output = str(project_root / "data" / "clinic-results" / "place_data_doctors.csv")

    count = convert(args.xlsx, args.sheet, args.output)
    print(f"Converted {count} hospitals from '{args.sheet}' -> {args.output}")


if __name__ == "__main__":
    main()
